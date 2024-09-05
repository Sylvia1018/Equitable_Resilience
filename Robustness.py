from OD_trip_request_for_TCS2011 import *
from datetime import datetime, timedelta
from numpy import nan
import numpy as np
import math
import networkx as nx
import openpyxl
from collections import defaultdict
import copy
import random


def time_range(x):
    for t in range(24):
        if datetime.time(t, 0) <= x <= datetime.time(t, 59):
            # return t
            if t in [7, 8, 9]:
                return 1  # AM peak
            elif t in [17, 18, 19]:
                return 2  # PM peak
            else:
                return 0  # non-peak
    print('time filter error:', x)


def age_range(x):
    # if 1 < x < 100:
    #     i = (x - 8) / 5.0 + 1
    #     if i < 1:
    #         return 0
    #     else:
    #         return int(math.floor(i))
    if 0 < x <= 14:
        return 0  # children
    elif 14 < x <= 24:
        return 1  # senior students / young adults
    elif 24 < x <= 34:
        return 2  # young working-age group
    elif 34 < x <= 44:
        return 3  #
    elif 44 < x <= 54:
        return 4  #
    elif 54 < x <= 64:
        return 5  #
    elif 64 < x:
        return 6  # retired / elderly

    else:
        print('age filter error:', x)


df = pd.read_csv('HK metro Space L 2011 reordered.csv', index_col=0)
station_index = pd.read_csv('OD matrix station index.csv')

trips = pd.read_csv('trips_df.csv')
trips['age_group'] = trips['age'].map(age_range)

station_index = pd.read_csv('OD matrix station index.csv')

# 合并数据集，将站台名称替换为 code reference
merged_df = trips.merge(station_index, left_on='origin', right_on='station', how='left')

merged_df = merged_df.rename(columns={'index': 'origin_index'})

merged_df_des = merged_df.merge(station_index, left_on='destination', right_on='station', how='left')

merged_df_des = merged_df_des.rename(columns={'index': 'destination_index'})

trip_index = merged_df_des[['time','sex','age','industry','income', 'origin','destination','origin_index','destination_index','age_group']]

df.index = range(len(df))
df.columns = range(len(df))

attr_group_dic = {'time': [[] for i in range(3)], 'sex': [[] for i in range(2)],
                'age': [[] for i in range(7)], 'industry': [[] for i in range(20)],
                'income': [[] for i in range(20)]
                 }

xlsx_file = openpyxl.load_workbook('group_data/age.xlsx')

data_dict = {}
for sheet_name in xlsx_file.sheetnames:
    sheet = xlsx_file[sheet_name]
    sheet_data = []
    for row in sheet.iter_rows(values_only=True, min_row=2):
        sheet_data.append(list(row[1:]))
    #             sheet_data[row[0]] = row[1:]  # 假设第一列是键，后面的列是值
    data_dict[sheet_name] = sheet_data
attr_group_dic['age'] = data_dict

xlsx_file = openpyxl.load_workbook('group_data/sex.xlsx')
data_dict = {}
for sheet_name in xlsx_file.sheetnames:
    sheet = xlsx_file[sheet_name]
    sheet_data = []
    for row in sheet.iter_rows(values_only=True, min_row=2):
        sheet_data.append(list(row[1:]))
#             sheet_data[row[0]] = row[1:]  # 假设第一列是键，后面的列是值
    data_dict[sheet_name] = sheet_data
attr_group_dic['sex'] = data_dict

def check_connected(G, data):
    return nx.has_path(G, data[-3], data[-2])


def get_network():
    G = nx.DiGraph()

    for node1 in df.index:
        for node2 in df.columns:
            if df.loc[node1, node2] == 1:
                G.add_edge(node1, node2)

    return G


def calculate_gini(data):
    #     data = np.array(list(data.values()))
    data = data.values()

    # 先将数据排序
    data_sorted = np.sort(list(data))

    # 计算累积相对频率
    total = sum(data_sorted)

    # 计算各类别频数累积占比
    cumulative_sum = np.cumsum(data_sorted)
    cumulative_percentage = cumulative_sum / total

    # 计算GINI指数
    gini = 1 - 2 * np.sum(cumulative_percentage) / len(data_sorted) + 1 / len(data_sorted)

    return gini


def get_edge_dict(G, update=True):
    if update:
        edge_dict = defaultdict(list)
        for edge in G.edges():
            x, y = edge[0], edge[1]
            edge_dict[x].append(y)

    return edge_dict


def o_d_matrix(data):
    record_matrix = pd.DataFrame([[0 for _ in range(len(station_index))] for _ in range(len(station_index))])

    record_matrix.columns = range(len(df.columns))

    record_matrix.index = range(len(df.columns))

    for record in data:
        ori = record[-3]
        des = record[-2]

        record_matrix[ori][des] += 1

    return record_matrix

def reachable_nodes(G,origin, edge_dict=None):
    current_node = origin
    visited = []
    path = [current_node]
    if edge_dict is None:
        edge_dict = get_edge_dict(G)
    while path:
        current_node = path.pop(0)
        visited.append(current_node)
        destinations = edge_dict[current_node]
        for next_node in destinations:
            if next_node not in visited and next_node not in path:
                path.append(next_node)
    return visited

trip_index_h = group_data_by_time_interval(trip_index.values, 60)


whole_bc_h = weight_betweenness(trip_index.values.tolist())

del trip_index_h[0], trip_index_h[1], trip_index_h[2], trip_index_h[3], trip_index_h[4], trip_index_h[5]

tmp_res = []
for value in trip_index_h.values():
    tmp_res.append(calculate_gini(weight_betweenness(value)))

# 使用map()函数将字典中的key替换为DataFrame中的另一列的值
station_index['bc'] = station_index['index'].map(wbc)
station_index = station_index.drop(['bc'], axis =1)

wbc = weight_betweenness(trip_index.values)


def simulation(total_simulations, data):
    #     total_simulations = 10
    integral_values = []
    remaining_traffic = []
    removed_site = []

    for _ in range(total_simulations):
        #         df = pd.DataFrame(data.copy())  # 创建一个副本以确保每次模拟从相同数据开始

        G = get_network()

        temp = copy.deepcopy(G)

        node_list, n0 = list(G.nodes()), G.number_of_nodes()

        remaining_traffic_ratios = []
        removed_site_ratios = []

        total_nodes = len(G.nodes())
        removed_nodes = []  # 已经被移除的节点
        flow_matrix = o_d_matrix(data)
        total_flow = flow_matrix.values.sum()

        step = 0
        # 模拟随机移除节点的过程
        while node_list:
            # 随机选择一个节点
            node_to_remove = random.choice(list(temp.nodes()))

            # 移除节点
            temp.remove_node(node_to_remove)
            removed_nodes.append(node_to_remove)
            node_list = list(temp.nodes())

            if node_list:
                curr_num = 0
                for origin in node_list:
                    reached = reachable_nodes(temp, origin)
                    for destination in reached:
                        curr_num += flow_matrix.loc[origin, destination]

            #             for record in data:
            #                 if G.has_node(record[-3]) and G.has_node(record[-2]) and check_connected(G, record):
            #                     curr_num+= 1

            remaining_traffic_ratios.append(curr_num / total_flow)
            removed_site_ratios.append(1 - len(temp.nodes()) / total_nodes)

        # 计算曲线的积分并记录
        integral = np.trapz(remaining_traffic_ratios, removed_site_ratios)
        #         plt.plot(remaining_traffic_ratios, removed_site_ratios)
        #         plt.show()
        integral_values.append(integral)
        remaining_traffic.append(remaining_traffic_ratios)
        removed_site.append(removed_site_ratios)

    # 统计所有模拟的积分结果
    mean_integral = np.mean(integral_values)
    std_deviation = np.std(integral_values)
    mean_remaining = [sum(col) / len(col) for col in zip(*remaining_traffic)]
    mean_removed = [sum(col) / len(col) for col in zip(*removed_site)]

    # 打印模拟结果统计
    print("Mean Integral over", total_simulations, "simulations:", mean_integral)
    print("Standard Deviation of Integral:", std_deviation)
    print("Average remaining is", mean_remaining)
    print("Average removed is", mean_removed)

    return [mean_integral, std_deviation, mean_remaining, mean_removed]

robustness_random = {'sex': [], 'age': [], 'industry': [], 'income': [], 'income_level': []}


for key, ls in tqdm(attr_group_dic['age'].items()):
    print(key)
    robustness_random['age'].append([simulation(500, ls)])

# for key, ls in tqdm(attr_group_dic['industry'].items()):
#     robustness_random['industry'].append([simulation(1000,ls)])

for key, ls in tqdm(attr_group_dic['sex'].items()):
    print(key)
    robustness_random['sex'].append([simulation(500, ls)])


def time_to_hours(time):
    # 将时间字符串转换为datetime对象
    time_datetime = datetime.strptime(time, '%H:%M:%S')  # 根据实际格式进行修改

    # 计算时间相对于午夜的小时数
    hours = time_datetime.hour + time_datetime.minute / 60 + time_datetime.second / 3600

    return hours


def group_data_by_time_interval(data, time_interval_minutes):
    # 将时间间隔转换为小时单位
    time_interval_hours = time_interval_minutes / 60

    # 创建一个字典，用于存储按时间分组的数据
    grouped_data = {}

    for item in data:
        # 获取数据行中的时间值

        time_value = item[0]

        # 将时间字符串转换为小时表示
        time_hours = time_to_hours(time_value)

        # 计算数据应该属于哪个时间间隔
        time_group = int(time_hours / time_interval_hours)

        # 如果时间间隔不存在于字典中，创建一个空列表
        if time_group not in grouped_data:
            grouped_data[time_group] = []

        # 将数据添加到对应时间间隔的列表中
        grouped_data[time_group].append(item)

    # 将字典键按时间顺序排序
    sorted_grouped_data = dict(sorted(grouped_data.items()))

    return sorted_grouped_data

def sort_BC(BC):
    return sorted(BC.items(), key=lambda x: x[1], )


def weight_betweenness(data):
    G = get_network()
    node_list, queue = list(G.nodes), list(G.nodes)
    flow_matrix = o_d_matrix(data)
    node_flow_centrality = {}
    total_flow = flow_matrix.values.sum()
    sp = {}

    while queue:
        v = queue.pop(0)
        od_flow_v = 0
        for e1 in node_list:
            for e2 in node_list:

                od_flow = flow_matrix.loc[e1, e2]
                if e1 != e2 and v not in [e1, e2] and od_flow > 0:
                    if (e1, e2) not in sp.keys():
                        try:
                            paths = list(nx.all_shortest_paths(G, e1, e2))
                            sp[e1, e2] = paths
                        except:
                            paths = None
                            sp[e1, e2] = paths
                    else:
                        paths = sp[e1, e2]
                    if paths:
                        nosp = len(paths)
                        for path in paths:

                            if v in path:
                                od_flow_v += od_flow / nosp

        if total_flow == 0:
            node_flow_centrality[v] = 0
        else:
            node_flow_centrality[v] = od_flow_v / total_flow
    return node_flow_centrality


def time_range(x):
    for t in range(24):
        if datetime.time(t, 0) <= x <= datetime.time(t, 59):
            # return t
            if t in [7, 8, 9]:
                return 1  # AM peak
            elif t in [17, 18, 19]:
                return 2  # PM peak
            else:
                return 0  # non-peak
    print('time filter error:', x)


def age_range(x):
    # if 1 < x < 100:
    #     i = (x - 8) / 5.0 + 1
    #     if i < 1:
    #         return 0
    #     else:
    #         return int(math.floor(i))
    if 0 < x <= 14:
        return 0  # children
    elif 14 < x <= 24:
        return 1  # senior students / young adults
    elif 24 < x <= 34:
        return 2  # young working-age group
    elif 34 < x <= 44:
        return 3  #
    elif 44 < x <= 54:
        return 4  #
    elif 54 < x <= 64:
        return 5  #
    elif 64 < x:
        return 6  # retired / elderly

    else:
        print('age filter error:', x)


# previous version of robustness analysis
def simulation_order(order, data):
    # order: Station&BC
    # data:station&count

    integral_values = []

    #     order = weight_matrix(order)
    #     order = order.rename(index=station_dic)
    #     data = data.apply(count_non_null, axis =1)
    order = sort_BC(order)

    #     for _ in range(total_simulations):

    G = get_network()

    node_list, n0 = list(G.nodes()), G.number_of_nodes()

    remaining_traffic_ratios = []
    removed_site_ratios = []

    total_nodes = len(G.nodes())
    removed_nodes = []  # 已经被移除的节点
    flow_matrix = o_d_matrix(data)

    total_flow = flow_matrix.values.sum()

    step = 0

    # 模拟随机移除节点的过程
    while node_list:
        # 随机选择一个节点
        node_to_remove = order[-1][0]

        order = order[:-1]
        #         node_to_remove = random.choice(list(G.nodes()))

        # 移除节点
        G.remove_node(node_to_remove)
        removed_nodes.append(node_to_remove)

        node_list = list(G.nodes())

        if node_list:
            curr_num = 0
            for origin in node_list:
                reached = reachable_nodes(G, origin)
                for destination in reached:
                    curr_num += flow_matrix.loc[origin, destination]

        remaining_traffic_ratios.append(curr_num / total_flow)
        removed_site_ratios.append(1 - len(G.nodes()) / total_nodes)

    #     # 模拟，每次移除一个站点
    #         while index!= len(order):
    #         # 随机选择一个站点并移除
    # #             site_to_remove = np.random.choice(df.index)
    #             if order.index[index] in df.index:

    # #                 print(order.index[index])
    #                 df.drop(index= order.index[index], inplace=True)

    #         # 计算当前剩余站点的总人流量
    #                 total_remaining_traffic = df.sum().sum()

    #         # 计算剩余人流量占比和移除站点占比
    #                 remaining_traffic_ratio = total_remaining_traffic / total_initial_traffic
    #                 removed_site_ratio = 1 - len(df.index) / len(data)

    #         # 存储数据
    #                 remaining_traffic_ratios.append(remaining_traffic_ratio)
    #                 removed_site_ratios.append(removed_site_ratio)
    # #             else: print('--------'+ order.index[index])
    #             index +=1

    # 计算曲线的积分并记录
    integral = np.trapz(remaining_traffic_ratios, removed_site_ratios)
    integral_values.append(integral)
    #     plt.plot(remaining_traffic_ratios, removed_site_ratios)
    #     plt.show()

    # 统计所有模拟的积分结果
    mean_integral = np.mean(integral_values)
    std_deviation = np.std(integral_values)

    # 打印模拟结果统计
    print("simulations:", mean_integral)
    #     print("Standard Deviation of Integral:", std_deviation)

    return mean_integral, remaining_traffic_ratios, removed_site_ratios

whole_bc = weight_betweenness(trip_index.values.tolist())


# previous version of robustness analysis
def simulation_NDorder(data):
    # order: Station&BC
    # data:station&count

    integral_values = []

    #     order = weight_matrix(order)
    #     order = order.rename(index=station_dic)
    #     data = data.apply(count_non_null, axis =1)

    #     for _ in range(total_simulations):

    G = get_network()

    node_list, n0 = list(G.nodes()), G.number_of_nodes()

    sorted_degrees = sorted(dict(G.degree()).items(), key=lambda x: x[1], reverse=True)

    # 提取排序后的节点列表
    order = [node for node, _ in sorted_degrees]

    remaining_traffic_ratios = []
    removed_site_ratios = []

    total_nodes = len(G.nodes())
    removed_nodes = []  # 已经被移除的节点
    flow_matrix = o_d_matrix(data)

    total_flow = flow_matrix.values.sum()

    step = 0

    # 模拟随机移除节点的过程
    while node_list:

        node_to_remove = order[0]

        order = order[1:]
        #         node_to_remove = random.choice(list(G.nodes()))

        # 移除节点
        G.remove_node(node_to_remove)
        removed_nodes.append(node_to_remove)

        node_list = list(G.nodes())

        if node_list:
            curr_num = 0
            for origin in node_list:
                reached = reachable_nodes(G, origin)
                for destination in reached:
                    curr_num += flow_matrix.loc[origin, destination]

        remaining_traffic_ratios.append(curr_num / total_flow)
        removed_site_ratios.append(1 - len(G.nodes()) / total_nodes)

    #     # 模拟，每次移除一个站点
    #         while index!= len(order):
    #         # 随机选择一个站点并移除
    # #             site_to_remove = np.random.choice(df.index)
    #             if order.index[index] in df.index:

    # #                 print(order.index[index])
    #                 df.drop(index= order.index[index], inplace=True)

    #         # 计算当前剩余站点的总人流量
    #                 total_remaining_traffic = df.sum().sum()

    #         # 计算剩余人流量占比和移除站点占比
    #                 remaining_traffic_ratio = total_remaining_traffic / total_initial_traffic
    #                 removed_site_ratio = 1 - len(df.index) / len(data)

    #         # 存储数据
    #                 remaining_traffic_ratios.append(remaining_traffic_ratio)
    #                 removed_site_ratios.append(removed_site_ratio)
    # #             else: print('--------'+ order.index[index])
    #             index +=1

    # 计算曲线的积分并记录
    integral = np.trapz(remaining_traffic_ratios, removed_site_ratios)
    integral_values.append(integral)
    #     plt.plot(remaining_traffic_ratios, removed_site_ratios)
    #     plt.show()

    # 统计所有模拟的积分结果
    mean_integral = np.mean(integral_values)
    std_deviation = np.std(integral_values)

    # 打印模拟结果统计
    print("simulations:", mean_integral)
    #     print("Standard Deviation of Integral:", std_deviation)

    return mean_integral, remaining_traffic_ratios, removed_site_ratios


# previous version of robustness analysis
def simulation_BCorder(data):
    # order: Station&BC
    # data:station&count

    integral_values = []

    #     order = weight_matrix(order)
    #     order = order.rename(index=station_dic)
    #     data = data.apply(count_non_null, axis =1)

    #     for _ in range(total_simulations):

    G = get_network()

    node_list, n0 = list(G.nodes()), G.number_of_nodes()

    sorted_degrees = sorted(nx.betweenness_centrality(G).items(), key=lambda x: x[1], reverse=True)

    # 提取排序后的节点列表
    order = [node for node, _ in sorted_degrees]

    remaining_traffic_ratios = []
    removed_site_ratios = []

    total_nodes = len(G.nodes())
    removed_nodes = []  # 已经被移除的节点
    flow_matrix = o_d_matrix(data)

    total_flow = flow_matrix.values.sum()

    step = 0

    # 模拟随机移除节点的过程
    while node_list:

        node_to_remove = order[0]

        order = order[1:]
        #         node_to_remove = random.choice(list(G.nodes()))

        # 移除节点
        G.remove_node(node_to_remove)
        removed_nodes.append(node_to_remove)

        node_list = list(G.nodes())

        if node_list:
            curr_num = 0
            for origin in node_list:
                reached = reachable_nodes(G, origin)
                for destination in reached:
                    curr_num += flow_matrix.loc[origin, destination]

        remaining_traffic_ratios.append(curr_num / total_flow)
        removed_site_ratios.append(1 - len(G.nodes()) / total_nodes)

    #     # 模拟，每次移除一个站点
    #         while index!= len(order):
    #         # 随机选择一个站点并移除
    # #             site_to_remove = np.random.choice(df.index)
    #             if order.index[index] in df.index:

    # #                 print(order.index[index])
    #                 df.drop(index= order.index[index], inplace=True)

    #         # 计算当前剩余站点的总人流量
    #                 total_remaining_traffic = df.sum().sum()

    #         # 计算剩余人流量占比和移除站点占比
    #                 remaining_traffic_ratio = total_remaining_traffic / total_initial_traffic
    #                 removed_site_ratio = 1 - len(df.index) / len(data)

    #         # 存储数据
    #                 remaining_traffic_ratios.append(remaining_traffic_ratio)
    #                 removed_site_ratios.append(removed_site_ratio)
    # #             else: print('--------'+ order.index[index])
    #             index +=1

    # 计算曲线的积分并记录
    integral = np.trapz(remaining_traffic_ratios, removed_site_ratios)
    integral_values.append(integral)
    #     plt.plot(remaining_traffic_ratios, removed_site_ratios)
    #     plt.show()

    # 统计所有模拟的积分结果
    mean_integral = np.mean(integral_values)
    std_deviation = np.std(integral_values)

    # 打印模拟结果统计
    print("simulations:", mean_integral)
    #     print("Standard Deviation of Integral:", std_deviation)

    return mean_integral, remaining_traffic_ratios, removed_site_ratios


# nd_rec = []
# for value in tqdm(trip_index.values):

#     if value[0] >= '06:00:00':
#         nd_rec.append(value)

#     print(value)
selected_rows = trip_index[trip_index['time'] > '06:00:00']
group_item = group_data_by_time_interval(selected_rows.values, 60)

robustness_nd_1h = defaultdict(list)

for sub_key, sub_value in tqdm(group_item.items()):
    robustness_nd_1h[sub_key].append(simulation_NDorder(sub_value))

#     robustness_nd_whole_1h[key] = robustness_nd_1h
# nd_rec = []
# for value in tqdm(trip_index.values):

#     if value[0] >= '06:00:00':
#         nd_rec.append(value)

#     print(value)
selected_rows = trip_index[trip_index['time'] > '06:00:00']
group_item = group_data_by_time_interval(selected_rows.values, 60)

robustness_pbc_1h = defaultdict(list)

for sub_key, sub_value in tqdm(group_item.items()):
    robustness_pbc_1h[sub_key].append(simulation_BCorder(sub_value))

#     robustness_nd_whole_1h[key] = robustness_nd_1h

def simulation_order_group(order, data):
    simulation_dic = []

    for key, value in data.items():
        simulation_dic.append(simulation_order(order[key], data[key]))

    #
    return simulation_dic

attr_BC = {'sex': {}, 'age': {}, 'industry': {}, 'income': {}}


def simulation_ND_group(data):
    simulation_dic = {}

    for key, value in data.items():
        simulation_dic[key] = simulation_NDorder(data[key])

    #
    return simulation_dic

robustness_ND['sex'] = simulation_ND_group(attr_group_dic['sex'])
robustness_ND['age'] = simulation_ND_group(attr_group_dic['age'])
robustness_ND['income'] = simulation_ND_group(attr_group_dic['income_level'])

trip_index_h = group_data_by_time_interval(trip_index.values, 60)


# whole_bc_h = weight_betweenness(trip_index.values.tolist())

del trip_index_h[0],trip_index_h[1],trip_index_h[2],trip_index_h[3], trip_index_h[4],trip_index_h[5],

whole_bc_h = {}

for key, item in trip_index_h.items():
    whole_bc_h[key] = weight_betweenness(item)

whole_gini_h = calculate_gini_group(whole_bc_h.values())

whole_random_rb = {}
for key, value in tqdm(trip_index_h.items()):
    whole_random_rb[key] = simulation(500,trip_index_h[key])

robustness_bc_1h = {}
for key,value in whole_bc_h.items():
    robustness_bc_1h[key] = simulation_order(value, trip_index_h[key])


def groupdata_simulation_by_time(attr, time_interval_minutes, total_simulations):
    record = []
    for item in attr:
        group_item = group_data_by_time_interval(item, time_interval_minutes)
        record_time = []
        #     print(group_item.values)
        for lst in group_item.values():
            #         print(lst)
            record_time.append(simulation(lst, total_simulations))

        record.append(record_time)

    return record


def peak_hour_rb(time_filter, attr, rob_type=['random', 'bc'], pk_time=['AM', 'PM']):
    robustness_attr = {}

    for key, value in tqdm(attr_group_dic[attr].items()):
        value = [sub_list for sub_list in value if sub_list[0] >= time_filter[0] and sub_list[0] <= time_filter[1]]

        if rob_type == 'bc':
            robustness_attr[key + '_' + pk_time] = simulation_order(weight_betweenness(value), value)
        if rob_type == 'random':
            robustness_attr[key + '_' + pk_time] = simulation(200, value)

    return robustness_attr

ran_attr_a = {'age':{}, 'income':{}, 'sex':{}, 'industry':{}}
ran_attr_p = {'age':{}, 'income':{}, 'sex':{}, 'industry':{}}
bc_attr_a = {'age':{}, 'income':{}, 'sex':{}, 'industry':{}}
bc_attr_p = {'age':{}, 'income':{}, 'sex':{}, 'industry':{}}

for key in attr_group_dic.keys():
    if key!= 'time':
        bc_attr_a[key] = peak_hour_rb(['07:00:00', '10:00:00'], key, rob_type = 'bc', pk_time = 'AM')

for key in attr_group_dic.keys():
    if key!= 'time':
        ran_attr_p[key] = peak_hour_rb(['17:00:00', '20:00:00'], key, rob_type = 'random', pk_time = 'PM')

for key in attr_group_dic.keys():
    if key!= 'time':
        bc_attr_p[key] = peak_hour_rb(['17:00:00', '20:00:00'], key, rob_type = 'bc', pk_time = 'PM')


robustness_attr_nd_1h = {}
robustness_attr_pbc_1h = {}
for key, value in tqdm(attr_group_dic['age'].items()):
    value = [sub_list for sub_list in value if sub_list[0] >= '06:00:00']
    #     print(value)
    group_item = group_data_by_time_interval(value, 60)
    #     print(group_item)

    robustness_pbc_1h_age = defaultdict(list)

    for sub_key, sub_value in group_item.items():
        robustness_pbc_1h_age[sub_key].append(simulation_BCorder(sub_value))

    robustness_attr_pbc_1h[key] = robustness_pbc_1h_age

for key, value in tqdm(attr_group_dic['sex'].items()):
    value = [sub_list for sub_list in value if sub_list[0] >= '06:00:00']
    #     print(value)
    group_item = group_data_by_time_interval(value, 60)
    #     print(group_item)

    robustness_pbc_1h_sex = defaultdict(list)

    for sub_key, sub_value in group_item.items():
        robustness_pbc_1h_sex[sub_key].append(simulation_BCorder(sub_value))

    robustness_attr_pbc_1h[key] = robustness_pbc_1h_sex

for key, value in tqdm(attr_group_dic['income_level'].items()):
    value = [sub_list for sub_list in value if sub_list[0] >= '06:00:00']
    #     print(value)
    group_item = group_data_by_time_interval(value, 60)
    #     print(group_item)

    robustness_pbc_1h_income = defaultdict(list)

    for sub_key, sub_value in group_item.items():
        robustness_pbc_1h_income[sub_key].append(simulation_BCorder(sub_value))

    robustness_attr_pbc_1h[key] = robustness_pbc_1h_income